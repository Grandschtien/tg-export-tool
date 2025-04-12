from telethon import TelegramClient
from openpyxl import Workbook, load_workbook
import os
from tqdm import tqdm

api_id = os.getenv("api_id")
api_hash = os.getenv("api_hash")

client = TelegramClient('session_name', api_id, api_hash)

async def export_chat():
    print("Connecting to Telegram...")
    await client.start()
    print("Successfully connected!")

    chat = input("Enter the chat ID or username: ").strip()
    try:
        chat = int(chat)
    except ValueError:
        if chat.startswith('@'):
            chat = chat[1:]

    download_media = input("Do you want to download media files? (y/n): ").strip().lower() == 'y'
    media_folder = "downloaded_media"
    if download_media and not os.path.exists(media_folder):
        os.makedirs(media_folder)

    print(f"Exporting all messages from chat '{chat}'")
    try:
        chat_entity = await client.get_entity(chat)
    except Exception as e:
        print(f"Error: Could not find chat '{chat}'. Please check the ID or username.")
        return

    total_messages = await client.get_messages(chat_entity, limit=0)
    total_count = total_messages.total

    output_filename = "chat_history_with_users.xlsx"
    if os.path.exists(output_filename):
        wb = load_workbook(output_filename)
        ws = wb.active
        existing_ids = {}
        for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
            cell = row[0]
            if cell.value is not None:
                existing_ids[int(cell.value)] = cell.row
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Chat History"
        ws.append(["Message ID", "Date", "Full Name", "Username", "Message", "Media File"])
        existing_ids = {}

    message_count = 0
    with tqdm(desc="Exporting messages", unit="message", total=total_count, dynamic_ncols=True) as pbar:
        async for message in client.iter_messages(chat_entity, reverse=True):
            if not message.sender_id:
                continue

            sender = await message.get_sender()
            if hasattr(sender, 'first_name'):
                full_name = sender.first_name or ""
                if sender.last_name:
                    full_name += f" {sender.last_name}"
                username = sender.username or "No username"
            elif hasattr(sender, 'title'):
                full_name = sender.title
                username = "Channel/Group"
            else:
                full_name = "Unknown"
                username = "No username"

            media_path = ""
            if download_media and message.media:
                try:
                    media_path = await message.download_media(file=media_folder)
                except Exception as e:
                    print(f"Error downloading media: {e}")

            row_data = [
                message.id,
                message.date.strftime("%Y-%m-%d %H:%M:%S"),
                full_name,
                username,
                message.text or 'Media/Other content',
                media_path or ""
            ]

            if message.id in existing_ids:
                row_idx = existing_ids[message.id]
                for col_idx, value in enumerate(row_data, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            else:
                ws.append(row_data)

            message_count += 1
            pbar.update(1)

    wb.save(output_filename)
    print(f"Total messages processed: {message_count}")
    print(f"Excel file updated: '{output_filename}'")

with client:
    try:
        client.loop.run_until_complete(export_chat())
    except Exception as e:
        print(f"An error occurred: {e}")
